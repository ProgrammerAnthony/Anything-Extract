"""文档解析工具"""

import base64
from pathlib import Path
from typing import Any, Dict, List, Optional
import uuid

from services.qanything_parser_bridge import QAnythingParserBridge
from services.runtime_config_service import runtime_config_service
from utils.loaders import CSVLoader, JSONLoader, convert_markdown_to_langchaindoc
from utils.logging import document_logger

try:
    from langchain_community.document_loaders import (
        TextLoader,
        UnstructuredEmailLoader,
        UnstructuredFileLoader,
        UnstructuredPowerPointLoader,
        UnstructuredWordDocumentLoader,
    )
except ImportError as exc:
    raise RuntimeError("缺少 langchain_community 依赖，请先安装 requirements.txt") from exc


class DocumentParser:
    """文档解析器"""

    SUPPORTED_TYPES = {
        "pdf",
        "docx",
        "txt",
        "md",
        "csv",
        "json",
        "xlsx",
        "pptx",
        "eml",
        "jpg",
        "jpeg",
        "png",
    }

    def __init__(self) -> None:
        self.bridge = QAnythingParserBridge()

    async def parse(
        self,
        file_path: str,
        file_type: Optional[str] = None,
        parser_mode: Optional[str] = None,
    ) -> Dict[str, Any]:
        """解析文档，返回统一结构"""
        parsed_type = self._normalize_file_type(file_path, file_type)
        parser_config = runtime_config_service.get_parser_config()
        resolved_mode = self._resolve_parser_mode(parser_mode, parser_config.get("mode"))

        if parsed_type == "pdf":
            return await self._parse_pdf(file_path, parser_mode=resolved_mode, parser_config=parser_config)
        if parsed_type in {"jpg", "jpeg", "png"}:
            return await self._parse_image(
                file_path,
                file_type=parsed_type,
                parser_mode=resolved_mode,
                parser_config=parser_config,
            )

        parser = getattr(self, f"_parse_{parsed_type}", None)
        if parser is None:
            raise ValueError(f"不支持的文件类型: {parsed_type}")

        return await parser(file_path)

    def _resolve_parser_mode(self, parser_mode: Optional[str], default_mode: Optional[str]) -> str:
        normalized = (parser_mode or default_mode or "local").strip().lower()
        if normalized not in {"local", "server", "hybrid"}:
            return "local"
        return normalized

    def _normalize_file_type(self, file_path: str, file_type: Optional[str]) -> str:
        if file_type:
            normalized = file_type.lower().lstrip(".")
        else:
            normalized = Path(file_path).suffix.lower().lstrip(".")

        if normalized not in self.SUPPORTED_TYPES:
            raise ValueError(f"不支持的文件类型: {normalized}")
        return normalized

    def _build_result(
        self,
        file_path: str,
        pages: List[Dict[str, Any]],
        metadata: Optional[Dict[str, Any]] = None,
        title: Optional[str] = None,
    ) -> Dict[str, Any]:
        normalized_pages: List[Dict[str, Any]] = []
        for idx, page in enumerate(pages, start=1):
            content = (page.get("content") or "").strip()
            if not content:
                continue
            normalized_pages.append({"page_number": idx, "content": content})

        if not normalized_pages:
            normalized_pages = [{"page_number": 1, "content": ""}]

        return {
            "id": str(uuid.uuid4()),
            "pages": normalized_pages,
            "title": title or Path(file_path).stem,
            "metadata": {
                "page_count": len(normalized_pages),
                **(metadata or {}),
            },
        }

    def _docs_to_pages(self, docs: List[Any]) -> List[Dict[str, Any]]:
        pages = []
        for index, doc in enumerate(docs, start=1):
            pages.append(
                {
                    "page_number": index,
                    "content": (getattr(doc, "page_content", "") or "").strip(),
                }
            )
        return pages

    async def _parse_pdf(
        self,
        file_path: str,
        parser_mode: str = "local",
        parser_config: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        parser_config = parser_config or runtime_config_service.get_parser_config()
        enable_pdf_server = bool(parser_config.get("enable_pdf_parser_server", False))

        if parser_mode in {"server", "hybrid"}:
            if enable_pdf_server:
                try:
                    return await self._parse_pdf_via_server(file_path, parser_mode, parser_config)
                except Exception as exc:  # noqa: BLE001
                    if parser_mode == "server":
                        raise RuntimeError(f"PDF 解析服务调用失败: {exc}") from exc
                    document_logger.warning("PDF parser server failed, fallback to local parser: %s", exc)
            elif parser_mode == "server":
                raise ValueError("解析策略为 server，但未启用 PDF 解析服务")

        return await self._parse_pdf_local(file_path, parser_mode, parser_config)

    async def _parse_pdf_via_server(
        self,
        file_path: str,
        parser_mode: str,
        parser_config: Dict[str, Any],
    ) -> Dict[str, Any]:
        bridge_result = await self.bridge.call_pdf_parser(
            filename=file_path,
            save_dir=str(Path(file_path).parent),
            pdf_parser_server_url=str(parser_config.get("pdf_parser_server_url", "")).strip(),
        )

        markdown_file = bridge_result.get("markdown_file")
        if not markdown_file:
            raise RuntimeError("PDF 解析服务未返回 markdown_file")

        markdown_path = Path(markdown_file)
        if not markdown_path.exists():
            raise RuntimeError(f"PDF 解析服务返回的 markdown 文件不存在: {markdown_file}")

        docs = convert_markdown_to_langchaindoc(str(markdown_path))
        pages = []
        for idx, doc in enumerate(docs, start=1):
            title_lst = doc.metadata.get("title_lst", []) if hasattr(doc, "metadata") else []
            title_prefix = "\n".join(title_lst)
            content = (doc.page_content or "").strip()
            merged = f"{title_prefix}\n{content}".strip() if title_prefix else content
            pages.append({"page_number": idx, "content": merged})

        result = self._build_result(
            file_path=file_path,
            pages=pages,
            metadata={
                "parser_mode_requested": parser_mode,
                "parser_strategy": "pdf_parser_server",
                "parser_source": "server",
                "parser_server_url": bridge_result.get("endpoint"),
                "parser_server_elapsed_ms": bridge_result.get("elapsed_ms"),
                "pdf_markdown_file": str(markdown_path),
            },
        )
        return result

    async def _parse_pdf_local(
        self,
        file_path: str,
        parser_mode: str,
        parser_config: Dict[str, Any],
    ) -> Dict[str, Any]:
        try:
            import PyPDF2

            with open(file_path, "rb") as file:
                pdf_reader = PyPDF2.PdfReader(file)
                pages = []
                for page_num, page in enumerate(pdf_reader.pages, start=1):
                    pages.append({"page_number": page_num, "content": page.extract_text() or ""})

                metadata = pdf_reader.metadata or {}
                return self._build_result(
                    file_path=file_path,
                    pages=pages,
                    title=metadata.get("/Title") or Path(file_path).stem,
                    metadata={
                        "author": metadata.get("/Author", ""),
                        "parser_mode_requested": parser_mode,
                        "parser_strategy": "local_pypdf2",
                        "parser_source": "local",
                        "pdf_server_enabled": bool(parser_config.get("enable_pdf_parser_server", False)),
                    },
                )
        except Exception:
            pass

        try:
            import pdfplumber

            with pdfplumber.open(file_path) as pdf:
                pages = []
                for page_num, page in enumerate(pdf.pages, start=1):
                    pages.append({"page_number": page_num, "content": page.extract_text() or ""})
                return self._build_result(
                    file_path=file_path,
                    pages=pages,
                    metadata={
                        "parser_mode_requested": parser_mode,
                        "parser_strategy": "local_pdfplumber",
                        "parser_source": "local",
                        "pdf_server_enabled": bool(parser_config.get("enable_pdf_parser_server", False)),
                    },
                )
        except Exception:
            loader = UnstructuredFileLoader(file_path, mode="fast")
            docs = loader.load()
            return self._build_result(
                file_path=file_path,
                pages=self._docs_to_pages(docs),
                metadata={
                    "source": "unstructured_pdf_fallback",
                    "parser_mode_requested": parser_mode,
                    "parser_strategy": "local_unstructured_fallback",
                    "parser_source": "local",
                    "pdf_server_enabled": bool(parser_config.get("enable_pdf_parser_server", False)),
                },
            )

    async def _parse_docx(self, file_path: str) -> Dict[str, Any]:
        try:
            loader = UnstructuredWordDocumentLoader(file_path, mode="fast")
            docs = loader.load()
            return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))
        except Exception:
            from docx import Document as DocxDocument

            doc = DocxDocument(file_path)
            content = []
            for paragraph in doc.paragraphs:
                if paragraph.text.strip():
                    content.append(paragraph.text.strip())

            return self._build_result(
                file_path=file_path,
                pages=[{"page_number": 1, "content": "\n".join(content)}],
                title=doc.core_properties.title or Path(file_path).stem,
                metadata={
                    "author": doc.core_properties.author or "",
                    "paragraph_count": len(content),
                },
            )

    async def _parse_txt(self, file_path: str) -> Dict[str, Any]:
        loader = TextLoader(file_path, autodetect_encoding=True)
        docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_md(self, file_path: str) -> Dict[str, Any]:
        try:
            docs = convert_markdown_to_langchaindoc(file_path)
            pages = []
            for idx, doc in enumerate(docs, start=1):
                title_lst = doc.metadata.get("title_lst", []) if hasattr(doc, "metadata") else []
                title_prefix = "\n".join(title_lst)
                content = (doc.page_content or "").strip()
                merged = f"{title_prefix}\n{content}".strip() if title_prefix else content
                pages.append({"page_number": idx, "content": merged})
            return self._build_result(file_path=file_path, pages=pages)
        except Exception:
            loader = UnstructuredFileLoader(file_path, mode="fast")
            docs = loader.load()
            return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_csv(self, file_path: str) -> Dict[str, Any]:
        loader = CSVLoader(
            file_path,
            autodetect_encoding=True,
            csv_args={"delimiter": ",", "quotechar": '"'},
        )
        docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_json(self, file_path: str) -> Dict[str, Any]:
        loader = JSONLoader(file_path, autodetect_encoding=True)
        docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_xlsx(self, file_path: str) -> Dict[str, Any]:
        pages = self._xlsx_to_markdown_pages(file_path)
        if pages:
            return self._build_result(file_path=file_path, pages=pages)

        loader = UnstructuredFileLoader(file_path, mode="fast")
        docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_pptx(self, file_path: str) -> Dict[str, Any]:
        try:
            loader = UnstructuredPowerPointLoader(file_path, mode="fast")
            docs = loader.load()
        except Exception:
            loader = UnstructuredFileLoader(file_path, mode="fast")
            docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_eml(self, file_path: str) -> Dict[str, Any]:
        try:
            loader = UnstructuredEmailLoader(file_path, mode="fast")
            docs = loader.load()
        except Exception:
            loader = TextLoader(file_path, autodetect_encoding=True)
            docs = loader.load()
        return self._build_result(file_path=file_path, pages=self._docs_to_pages(docs))

    async def _parse_image(
        self,
        file_path: str,
        file_type: str,
        parser_mode: str = "local",
        parser_config: Optional[Dict[str, Any]] = None,
    ) -> Dict[str, Any]:
        parser_config = parser_config or runtime_config_service.get_parser_config()
        enable_ocr_server = bool(parser_config.get("enable_ocr_server", False))

        if parser_mode in {"server", "hybrid"}:
            if enable_ocr_server:
                try:
                    return await self._parse_image_via_server(file_path, file_type, parser_mode, parser_config)
                except Exception as exc:  # noqa: BLE001
                    if parser_mode == "server":
                        raise RuntimeError(f"OCR 服务调用失败: {exc}") from exc
                    document_logger.warning("OCR server failed, fallback to local parser: %s", exc)
            elif parser_mode == "server":
                raise ValueError("解析策略为 server，但未启用 OCR 服务")

        return await self._parse_image_local(file_path, file_type, parser_mode, parser_config)

    async def _parse_image_via_server(
        self,
        file_path: str,
        file_type: str,
        parser_mode: str,
        parser_config: Dict[str, Any],
    ) -> Dict[str, Any]:
        file_bytes = Path(file_path).read_bytes()
        img64 = base64.b64encode(file_bytes).decode("utf-8")

        bridge_result = await self.bridge.call_ocr_server(
            img64=img64,
            ocr_server_url=str(parser_config.get("ocr_server_url", "")).strip(),
        )

        lines = bridge_result.get("lines") or []
        if not lines:
            raise RuntimeError("OCR 服务未识别出文本")

        result = self._build_result(
            file_path=file_path,
            pages=[{"page_number": 1, "content": "\n".join(lines)}],
            metadata={
                "parser_mode_requested": parser_mode,
                "parser_strategy": "ocr_server",
                "parser_source": "server",
                "parser_server_url": bridge_result.get("endpoint"),
                "parser_server_elapsed_ms": bridge_result.get("elapsed_ms"),
                "file_type": file_type,
            },
        )
        return result

    async def _parse_image_local(
        self,
        file_path: str,
        file_type: str,
        parser_mode: str,
        parser_config: Dict[str, Any],
    ) -> Dict[str, Any]:
        try:
            loader = UnstructuredFileLoader(file_path, mode="fast")
            docs = loader.load()
            pages = self._docs_to_pages(docs)
            if any((page.get("content") or "").strip() for page in pages):
                return self._build_result(
                    file_path=file_path,
                    pages=pages,
                    metadata={
                        "parser_mode_requested": parser_mode,
                        "parser_strategy": "local_unstructured_image",
                        "parser_source": "local",
                        "ocr_server_enabled": bool(parser_config.get("enable_ocr_server", False)),
                        "file_type": file_type,
                    },
                )
        except Exception:
            pass

        return self._build_result(
            file_path=file_path,
            pages=[{"page_number": 1, "content": "图片已上传。当前未启用 OCR 服务，未提取到可检索文本。"}],
            metadata={
                "parser_mode_requested": parser_mode,
                "parser_strategy": "local_image_placeholder",
                "parser_source": "local",
                "ocr_server_enabled": bool(parser_config.get("enable_ocr_server", False)),
                "file_type": file_type,
            },
        )

    def _xlsx_to_markdown_pages(self, file_path: str) -> List[Dict[str, Any]]:
        try:
            import openpyxl
        except ImportError:
            return []

        def clean_cell_content(cell: Any) -> str:
            if cell is None:
                return ""
            return " ".join(str(cell).split())

        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        pages: List[Dict[str, Any]] = []

        for page_number, sheet_name in enumerate(workbook.sheetnames, start=1):
            sheet = workbook[sheet_name]
            rows = [[clean_cell_content(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
            non_empty_rows = [row for row in rows if any(cell != "" for cell in row)]
            if not non_empty_rows:
                continue

            max_cols = max(len(row) for row in non_empty_rows)
            lines = [f"# {sheet_name}", ""]
            for row_index, row in enumerate(non_empty_rows):
                padded_row = row + [""] * (max_cols - len(row))
                lines.append("| " + " | ".join(padded_row) + " |")
                if row_index == 0:
                    lines.append("|" + "|".join(["---"] * max_cols) + "|")
            pages.append({"page_number": page_number, "content": "\n".join(lines)})

        workbook.close()
        return pages
